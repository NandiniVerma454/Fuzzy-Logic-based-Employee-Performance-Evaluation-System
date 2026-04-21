import streamlit as st
import pandas as pd
import matplotlib.pyplot as plt
from fuzzy_logic import evaluate_performance
import io
from datetime import datetime

st.set_page_config(page_title="Employee Performance System", layout="wide")
st.markdown("""
<h1 style='color:#FFD700; text-align:center;'>
📈 Fuzzy Logic based Employee Performance Evaluation System 📉
</h1>
""", unsafe_allow_html=True)
st.markdown("""
<style>
/* Sidebar background */
section[data-testid="stSidebar"] {
    background: linear-gradient(180deg, #0f2027, #203a43, black);
    color: yellow;
}

/* Sidebar text */
section[data-testid="stSidebar"] * {
    color: yellow;
    font-size: 14px;
}

/* Sidebar headings */
section[data-testid="stSidebar"] h2 {
    color: white;
    font-size: 20px;
}

/* Optional: make divider lighter */
section[data-testid="stSidebar"] hr {
    border-color: rgba(255,255,255,0.2);
}
</style>
""", unsafe_allow_html=True)

st.sidebar.markdown("## 📌 INSTRUCTIONS")
st.sidebar.write("---")

st.sidebar.markdown("""
<div>

<p><span style="color:#FFD700; font-weight:bold;">1️⃣ Enter Employee Details</span><br>
<span style="color:#FFF8DC;">Fill in all required fields such as name, age, job role, and income.</span></p>

<p><span style="color:#FFD700; font-weight:bold;">2️⃣ Provide Performance Inputs</span><br>
<span style="color:#FFF8DC;">Enter values for attendance, work quality, teamwork, task completion, job involvement, and work-life balance.</span></p>

<p><span style="color:#FFD700; font-weight:bold;">3️⃣ Click Evaluate Performance</span><br>
<span style="color:#FFF8DC;">Press the Evaluate Performance button to process the data.</span></p>

<p><span style="color:#FFD700; font-weight:bold;">4️⃣ View Results</span><br>
<span style="color:#FFF8DC;">The system will display performance score and category.</span></p>

<p><span style="color:#FFD700; font-weight:bold;">5️⃣ Download Report 📥</span><br>
<span style="color:#FFF8DC;">Download the report for record keeping, comparison, and tracking improvements.</span></p>

</div>
""", unsafe_allow_html=True)
import base64

def set_bg(image_file):
    with open(image_file, "rb") as f:
        encoded = base64.b64encode(f.read()).decode()

    st.markdown(f"""
        <style>
        .stApp {{
            background-image: url("data:image/jpg;base64,{encoded}");
            background-size: cover;
            background-position: center;
            background-repeat: no-repeat;
            background-attachment: fixed;
        }}
        
        </style>
    """, unsafe_allow_html=True)
set_bg("img.jpg")

current_date = datetime.now().strftime("%Y-%m-%d")

# ===================== EMPLOYEE DETAILS =====================
st.markdown("## 👤 Employee Details")
st.markdown("Enter the following parameters to evaluate employee performance:")
col1, col2 = st.columns(2)

with col1:
    name = st.text_input("Employee Name")
    age = st.number_input("Age", 18, 60, 25)
    gender = st.selectbox("Gender", ["Male", "Female", "Other"])

with col2:
    marital = st.selectbox("Marital Status", ["Single", "Married", "Divorced"])
    job_role = st.text_input("Job Role")
    income = st.number_input("Monthly Income", 1000, 100000, 20000)
st.markdown("---")   
st.markdown("## 🧾 Employee Performance Input Panel")

col1, col2 = st.columns(2)

with col1:
    st.markdown("### 📊 Work Metrics")
    attendance = st.slider("Attendance (%)", 0, 100, 75)
    quality = st.slider("Work Quality (1-10)", 1, 10, 7)
    task = st.slider("Task Completion (1-10)", 1, 10, 7)

with col2:
    st.markdown("### 🤝 Behavioral Metrics")
    teamwork = st.slider("Teamwork (1-10)", 1, 10, 6)
    involvement = st.slider("Job Involvement (1-10)", 1, 10, 7)
    wlb = st.slider("Work-Life Balance (1-10)", 1, 10, 6)
# ===================== EVALUATION =====================
st.markdown("---")
st.markdown("## 🎯 Performance Evaluation")

if st.button("Evaluate Performance"):

    # ✅ Step 1: Calculate score
    score, level = evaluate_performance(
        attendance, quality, teamwork, task, involvement, wlb
    )

    # ✅ Step 2: Show result
    st.metric("📈 Performance Score", f"{score:.2f}")
    st.success(f"🌟 Performance Level: {level}")

    # ================= EXCEL REPORT =================
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment

    wb = Workbook()
    ws = wb.active

    bold = Font(bold=True)
    center = Alignment(horizontal="center")

    # Header
    ws.merge_cells('A1:B1')
    ws['A1'] = "ABC Corporation Pvt. Ltd."
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = center

    ws.merge_cells('A2:B2')
    ws['A2'] = "Employee Performance Report"
    ws['A2'].alignment = center

    row = 4

    # Employee Details
    ws[f"A{row}"] = "Employee Details"
    ws[f"A{row}"].font = bold

    details = [
        ("Name", name),
        ("Age", age),
        ("Gender", gender),
        ("Marital Status", marital),
        ("Job Role", job_role),
        ("Monthly Income", income),
    ]

    for key, value in details:
        row += 1
        ws[f"A{row}"] = key
        ws[f"B{row}"] = value

    # Performance Inputs
    row += 2
    ws[f"A{row}"] = "Performance Inputs"
    ws[f"A{row}"].font = bold

    inputs = [
        ("Attendance (%)", attendance),
        ("Work Quality", quality),
        ("Teamwork", teamwork),
        ("Task Completion", task),
        ("Job Involvement", involvement),
        ("Work-Life Balance", wlb),
    ]

    for key, value in inputs:
        row += 1
        ws[f"A{row}"] = key
        ws[f"B{row}"] = value

    # Final Result
    row += 2
    ws[f"A{row}"] = "Final Result"
    ws[f"A{row}"].font = bold

    row += 1
    ws[f"A{row}"] = "Performance Score"
    ws[f"B{row}"] = round(score, 2)

    row += 1
    ws[f"A{row}"] = "Performance Level"
    ws[f"B{row}"] = level

    # Save
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    st.download_button(
        label="⬇️ Download Excel Report",
        data=buffer,
        file_name=f"{name}_performance_report.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    st.markdown("---")
    # ===================== INPUT VISUALIZATION =====================
    st.subheader("📊 Performance Interpretation")

    import matplotlib.pyplot as plt

    fig, ax = plt.subplots(figsize=(8, 2))

    # Soft color regions
    ax.axvspan(0, 40, color='#f2a7a7', alpha=0.4)   # Soft red (Poor)
    ax.axvspan(40, 60, color='#f7e6a1', alpha=0.5)  # Soft yellow (Average)
    ax.axvspan(60, 80, color='#a8d5ba', alpha=0.5)  # Soft green (Good)
    ax.axvspan(80, 100, color='#a7c7e7', alpha=0.5) # Soft blue (Excellent)

    # Score line
    ax.axvline(score, color='black', linewidth=3)

    # Axis settings
    ax.set_xlim(0, 100)
    ax.set_yticks([])
    ax.set_xlabel("Performance Score")

    # Labels
    ax.text(20, 0.5, "Poor", ha='center', fontsize=11)
    ax.text(50, 0.5, "Average", ha='center', fontsize=11)
    ax.text(70, 0.5, "Good", ha='center', fontsize=11)
    ax.text(90, 0.5, "Excellent", ha='center', fontsize=11)

    st.pyplot(fig)